from __future__ import annotations

from dataclasses import dataclass, field
from datetime import date, datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, BinaryIO

from openpyxl import Workbook, load_workbook
from openpyxl.comments import Comment
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.orm import Session

from archive_workbench.catalog import ensure_project
from archive_workbench.catalog_management import (
    REGISTRATION_STATUSES,
    archival_field_rows,
    catalog_unit_rows,
    create_archival_unit,
    move_archival_unit,
    update_archival_unit,
)
from archive_workbench.contracts.decisions import ProjectDecisions
from archive_workbench.db.models import ArchivalUnit

TEMPLATE_SCHEMA_VERSION = "1.0"
TEMPLATE_SHEETS = ("INSTRUCCIONES", "ESTRUCTURA", "CATALOGO", "LISTAS")
TEMPLATE_ACTIONS = ("", "crear", "actualizar", "omitir")
BOOLEAN_VALUES = ("", "sí", "no")
FIELD_STATES = ("", "provided", "no_information", "not_applicable", "pending")

_BASE_COLUMNS: tuple[tuple[str, str, str], ...] = (
    ("local_id", "ID local", "Identificador estable dentro de la plantilla. No puede repetirse."),
    (
        "unit_id",
        "ID de unidad existente",
        "Identificador interno exportado por Archive Workbench. Dejalo vacío para crear una unidad.",
    ),
    (
        "parent_local_id",
        "ID padre local",
        "ID local de la unidad padre incluida en esta misma plantilla.",
    ),
    (
        "parent_unit_id",
        "ID padre existente",
        "Identificador interno de una unidad padre que ya existe en el proyecto y no está en la plantilla.",
    ),
    (
        "action",
        "Acción",
        "Vacío aplica modo automático: crear si no hay ID existente; actualizar si lo hay.",
    ),
    ("level_key", "Nivel", "Clave de nivel archivístico definida en la hoja ESTRUCTURA."),
    ("reference_code", "Código de referencia", "Código de referencia archivístico, si existe."),
    ("title", "Título", "Título principal de la unidad archivística. Es obligatorio."),
    (
        "registration_status",
        "Estado de registro",
        "incomplete, provisional o complete. Un registro complete requiere confirmación explícita.",
    ),
    (
        "completion_confirmed",
        "Descripción completa confirmada",
        "Usá sí o no. La confirmación no se infiere automáticamente.",
    ),
    (
        "source_url",
        "URL de procedencia",
        "URL de la fuente utilizada para describir esta fila. Se conserva como nota de campo.",
    ),
    (
        "source_note",
        "Nota de procedencia",
        "Aclaraciones sobre la fuente, vacíos, decisiones técnicas o aspectos que requieren revisión.",
    ),
)


@dataclass(slots=True)
class CatalogTemplateIssue:
    severity: str
    code: str
    sheet: str
    row: int | None
    column: str | None
    message: str


@dataclass(slots=True)
class CatalogTemplateRow:
    row_number: int
    local_id: str
    unit_id: str | None
    parent_local_id: str | None
    parent_unit_id: str | None
    action: str
    level_key: str
    reference_code: str | None
    title: str
    registration_status: str
    completion_confirmed: bool
    source_url: str | None
    source_note: str | None
    field_values: dict[str, dict[str, Any]] = field(default_factory=dict)


@dataclass(slots=True)
class CatalogTemplateReport:
    schema_version: str
    template_name: str
    target_project_id: str | None
    valid: bool
    rows: list[CatalogTemplateRow]
    structure_parent_keys: dict[str, tuple[str, ...]]
    issues: list[CatalogTemplateIssue]
    create_count: int
    update_count: int
    skip_count: int

    @property
    def error_count(self) -> int:
        return sum(item.severity == "error" for item in self.issues)

    @property
    def warning_count(self) -> int:
        return sum(item.severity == "warning" for item in self.issues)

    def as_dict(self) -> dict[str, Any]:
        return {
            "schema_version": self.schema_version,
            "template_name": self.template_name,
            "target_project_id": self.target_project_id,
            "valid": self.valid,
            "rows": len(self.rows),
            "create_count": self.create_count,
            "update_count": self.update_count,
            "skip_count": self.skip_count,
            "error_count": self.error_count,
            "warning_count": self.warning_count,
            "structure_parent_keys": {
                key: list(value) for key, value in sorted(self.structure_parent_keys.items())
            },
            "issues": [
                {
                    "severity": item.severity,
                    "code": item.code,
                    "sheet": item.sheet,
                    "row": item.row,
                    "column": item.column,
                    "message": item.message,
                }
                for item in self.issues
            ],
        }


@dataclass(slots=True)
class CatalogTemplateApplyResult:
    created: int
    updated: int
    moved: int
    unchanged: int
    skipped: int
    local_to_unit_id: dict[str, str]


@dataclass(slots=True)
class _ParsedTemplate:
    schema_version: str
    template_name: str
    target_project_id: str | None
    structure_parent_keys: dict[str, tuple[str, ...]]
    rows: list[CatalogTemplateRow]
    parse_issues: list[CatalogTemplateIssue]


def _utc_now_iso() -> str:
    return datetime.now(timezone.utc).replace(microsecond=0).isoformat()


def _clean_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _split_keys(value: Any) -> tuple[str, ...]:
    text = _clean_text(value)
    if not text:
        return ()
    normalized = text.replace("\n", ",").replace(";", ",")
    return tuple(item.strip() for item in normalized.split(",") if item.strip())


def _parse_bool(value: Any) -> bool | None:
    if isinstance(value, bool):
        return value
    text = _clean_text(value).casefold()
    if not text:
        return False
    if text in {"sí", "si", "true", "1", "x", "yes"}:
        return True
    if text in {"no", "false", "0"}:
        return False
    return None


def _column_key(header: Any) -> str | None:
    text = _clean_text(header)
    if not text:
        return None
    if "[" in text and text.endswith("]"):
        return text.rsplit("[", 1)[1][:-1].strip()
    return text


def _field_columns(decisions: ProjectDecisions) -> list[tuple[str, str, str]]:
    columns: list[tuple[str, str, str]] = []
    for item in decisions.descriptive_fields:
        if not item.enabled or item.key == "reference_code":
            continue
        applies = ", ".join(item.applies_to_levels)
        examples = "; ".join(item.examples) if item.examples else "Sin ejemplo configurado"
        description = (
            f"Tipo: {item.data_type}. Niveles: {applies}. "
            f"Repetible: {'sí' if item.repeatable else 'no'}. Ejemplos: {examples}."
        )
        columns.extend(
            [
                (f"field:{item.key}", item.label, description),
                (
                    f"field_state:{item.key}",
                    f"Estado · {item.label}",
                    "Vacío usa provided si hay valor y pending si no lo hay. También admite "
                    "no_information y not_applicable.",
                ),
                (
                    f"field_note:{item.key}",
                    f"Nota · {item.label}",
                    "Nota de fuente o aclaración específica de este campo.",
                ),
            ]
        )
    return columns


def _header_text(label: str, key: str) -> str:
    return f"{label} [{key}]"


def _style_workbook(workbook: Workbook) -> None:
    workbook.calculation.fullCalcOnLoad = True
    workbook.calculation.forceFullCalc = True


def _style_header(sheet, row: int, *, fill: str = "1F4E78") -> None:
    thin = Side(style="thin", color="B8C4CE")
    for cell in sheet[row]:
        if cell.value is None:
            continue
        cell.fill = PatternFill("solid", fgColor=fill)
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin)
    sheet.row_dimensions[row].height = 34


def _set_widths(sheet, widths: dict[int, float], default: float = 18) -> None:
    for index in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(index)].width = widths.get(index, default)


def _add_catalog_validations(sheet, headers: dict[str, int], lists_sheet, max_row: int = 5000) -> None:
    formulas = {
        "action": "=LISTAS!$B$2:$B$5",
        "level_key": f"=LISTAS!$A$2:$A${max(2, lists_sheet.max_row)}",
        "registration_status": "=LISTAS!$C$2:$C$4",
        "completion_confirmed": "=LISTAS!$D$2:$D$4",
    }
    for key, formula in formulas.items():
        column = headers.get(key)
        if column is None:
            continue
        validation = DataValidation(type="list", formula1=formula, allow_blank=True)
        validation.error = "Elegí un valor de la lista configurada."
        validation.errorTitle = "Valor no permitido"
        validation.prompt = "Usá uno de los valores de la lista."
        validation.promptTitle = "Valor controlado"
        sheet.add_data_validation(validation)
        validation.add(f"{get_column_letter(column)}2:{get_column_letter(column)}{max_row}")

    for key, column in headers.items():
        if not key.startswith("field_state:"):
            continue
        validation = DataValidation(type="list", formula1="=LISTAS!$E$2:$E$6", allow_blank=True)
        validation.error = "Elegí un estado de campo válido."
        validation.errorTitle = "Estado no permitido"
        sheet.add_data_validation(validation)
        validation.add(f"{get_column_letter(column)}2:{get_column_letter(column)}{max_row}")


def export_catalog_template_bytes(
    session: Session | None,
    *,
    decisions: ProjectDecisions,
    project_id: str,
    include_catalog: bool = False,
    template_name: str = "Plantilla de catálogo",
    target_project_id: str | None = None,
    target_project_name: str | None = None,
    source_url: str | None = None,
    source_retrieved_at: str | None = None,
    source_note: str | None = None,
    structure_parent_overrides: dict[str, list[str] | tuple[str, ...]] | None = None,
    seed_rows: list[dict[str, Any]] | None = None,
) -> bytes:
    workbook = Workbook()
    instructions = workbook.active
    instructions.title = "INSTRUCCIONES"
    structure = workbook.create_sheet("ESTRUCTURA")
    catalog = workbook.create_sheet("CATALOGO")
    lists = workbook.create_sheet("LISTAS")
    _style_workbook(workbook)

    instructions.merge_cells("A1:H1")
    instructions["A1"] = template_name
    instructions["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    instructions["A1"].fill = PatternFill("solid", fgColor="17365D")
    instructions["A1"].alignment = Alignment(horizontal="left", vertical="center")
    instructions.row_dimensions[1].height = 30

    metadata = [
        ("schema_version", TEMPLATE_SCHEMA_VERSION),
        ("template_name", template_name),
        ("target_project_id", target_project_id if target_project_id is not None else project_id),
        ("project_name", target_project_name or decisions.project_name),
        ("generated_at", _utc_now_iso()),
        ("source_url", source_url or ""),
        ("source_retrieved_at", source_retrieved_at or ""),
        ("source_note", source_note or ""),
    ]
    instructions["A3"] = "Metadato"
    instructions["B3"] = "Valor"
    _style_header(instructions, 3)
    for row_index, (key, value) in enumerate(metadata, start=4):
        instructions.cell(row=row_index, column=1, value=key)
        instructions.cell(row=row_index, column=2, value=value)
        instructions.cell(row=row_index, column=1).font = Font(bold=True)
        instructions.cell(row=row_index, column=2).alignment = Alignment(wrap_text=True, vertical="top")

    start = 14
    instructions.cell(row=start, column=1, value="Cómo completar la plantilla")
    instructions.cell(row=start, column=1).font = Font(size=13, bold=True, color="17365D")
    guidance = [
        "Completá la hoja CATALOGO. Cada fila representa una unidad archivística.",
        "Usá ID local para vincular padres e hijos dentro de la misma plantilla.",
        "No edites los ID internos exportados, salvo que estés reparando una plantilla bajo supervisión.",
        "La hoja ESTRUCTURA define los padres permitidos. Puede restringir el proyecto, nunca ampliarlo.",
        "La hoja LISTAS es auxiliar y permanece oculta para sostener los desplegables; no hace falta editarla.",
        "Antes de importar se ejecuta una simulación completa. Ningún cambio se aplica si hay errores.",
        "Los valores repetibles se separan con saltos de línea dentro de la misma celda.",
        "Los datos ausentes deben quedar vacíos; no los completes por inferencia.",
    ]
    for offset, text in enumerate(guidance, start=1):
        instructions.cell(row=start + offset, column=1, value=f"{offset}.")
        instructions.cell(row=start + offset, column=2, value=text)
        instructions.cell(row=start + offset, column=2).alignment = Alignment(wrap_text=True)

    dictionary_row = start + len(guidance) + 3
    dictionary_headers = ["Hoja", "Columna", "Clave", "Descripción"]
    for column, value in enumerate(dictionary_headers, start=1):
        instructions.cell(row=dictionary_row, column=column, value=value)
    _style_header(instructions, dictionary_row)
    dictionary_entries: list[tuple[str, str, str, str]] = []
    for key, label, description in _BASE_COLUMNS + tuple(_field_columns(decisions)):
        dictionary_entries.append(("CATALOGO", label, key, description))
    dictionary_entries.extend(
        [
            ("ESTRUCTURA", "Padres del proyecto", "project_parent_keys", "Regla vigente en decisions.yaml."),
            (
                "ESTRUCTURA",
                "Padres permitidos por la plantilla",
                "template_parent_keys",
                "Restricción distribuible. Debe ser igual o más estricta que el proyecto.",
            ),
        ]
    )
    for row_index, values in enumerate(dictionary_entries, start=dictionary_row + 1):
        for column, value in enumerate(values, start=1):
            instructions.cell(row=row_index, column=column, value=value)
            instructions.cell(row=row_index, column=column).alignment = Alignment(
                vertical="top", wrap_text=True
            )
    instructions.freeze_panes = "A4"
    _set_widths(instructions, {1: 22, 2: 60, 3: 30, 4: 74})

    structure_headers = [
        ("level_key", "Clave de nivel"),
        ("label", "Etiqueta"),
        ("plural_label", "Etiqueta plural"),
        ("display_order", "Orden"),
        ("project_parent_keys", "Padres del proyecto"),
        ("template_parent_keys", "Padres permitidos por la plantilla"),
        ("required_fields", "Campos obligatorios"),
        ("optional", "Opcional"),
        ("enabled", "Habilitado"),
    ]
    for column, (key, label) in enumerate(structure_headers, start=1):
        structure.cell(row=1, column=column, value=_header_text(label, key))
    _style_header(structure, 1)
    override_map = structure_parent_overrides or {}
    levels = sorted(
        [item for item in decisions.archival_levels if item.enabled], key=lambda item: item.display_order
    )
    for row_index, level in enumerate(levels, start=2):
        template_parents = tuple(override_map.get(level.key, level.parent_keys))
        values = [
            level.key,
            level.label,
            level.plural_label,
            level.display_order,
            ", ".join(level.parent_keys),
            ", ".join(template_parents),
            ", ".join(level.required_fields),
            "sí" if level.optional else "no",
            "sí" if level.enabled else "no",
        ]
        for column, value in enumerate(values, start=1):
            structure.cell(row=row_index, column=column, value=value)
            structure.cell(row=row_index, column=column).alignment = Alignment(
                vertical="top", wrap_text=True
            )
    structure.freeze_panes = "A2"
    structure.auto_filter.ref = f"A1:I{max(2, structure.max_row)}"
    _set_widths(structure, {1: 20, 2: 24, 3: 24, 4: 10, 5: 36, 6: 42, 7: 30, 8: 12, 9: 12})

    lists.append(["Niveles", "Acciones", "Estados de registro", "Booleanos", "Estados de campo"])
    max_list_rows = max(len(levels), len(TEMPLATE_ACTIONS), len(REGISTRATION_STATUSES), len(BOOLEAN_VALUES), len(FIELD_STATES))
    for index in range(max_list_rows):
        lists.append(
            [
                levels[index].key if index < len(levels) else None,
                TEMPLATE_ACTIONS[index] if index < len(TEMPLATE_ACTIONS) else None,
                REGISTRATION_STATUSES[index] if index < len(REGISTRATION_STATUSES) else None,
                BOOLEAN_VALUES[index] if index < len(BOOLEAN_VALUES) else None,
                FIELD_STATES[index] if index < len(FIELD_STATES) else None,
            ]
        )
    _style_header(lists, 1, fill="5B9BD5")
    lists.sheet_state = "hidden"

    columns = list(_BASE_COLUMNS) + _field_columns(decisions)
    header_map: dict[str, int] = {}
    for column, (key, label, description) in enumerate(columns, start=1):
        header_map[key] = column
        cell = catalog.cell(row=1, column=column, value=_header_text(label, key))
        cell.comment = Comment(description, "Archive Workbench")
    _style_header(catalog, 1)
    catalog.freeze_panes = "A2"
    catalog.auto_filter.ref = f"A1:{get_column_letter(len(columns))}1"

    gray_fill = PatternFill("solid", fgColor="E7E6E6")
    for key in ("unit_id", "parent_unit_id"):
        column = header_map[key]
        for row in range(2, 5001):
            catalog.cell(row=row, column=column).fill = gray_fill

    rows_to_write: list[dict[str, Any]] = []
    if include_catalog:
        if session is None:
            raise ValueError("Se requiere una sesión para exportar el catálogo actual")
        units = catalog_unit_rows(session, project_id)
        local_by_unit = {row.id: f"unit_{row.id.replace('-', '')[:12]}" for row in units}
        for unit in units:
            field_rows = archival_field_rows(session, unit.id)
            by_key: dict[str, list[Any]] = {}
            for item in field_rows:
                by_key.setdefault(item.field_key, []).append(item)
            payload: dict[str, Any] = {
                "local_id": local_by_unit[unit.id],
                "unit_id": unit.id,
                "parent_local_id": local_by_unit.get(unit.parent_id or ""),
                "parent_unit_id": "",
                "action": "actualizar",
                "level_key": unit.level_key,
                "reference_code": unit.reference_code or "",
                "title": unit.title,
                "registration_status": unit.registration_status,
                "completion_confirmed": "sí" if unit.completion_confirmed else "no",
                "source_url": "",
                "source_note": "",
            }
            for definition in decisions.descriptive_fields:
                if not definition.enabled or definition.key == "reference_code":
                    continue
                values = by_key.get(definition.key, [])
                provided = [str(item.value) for item in values if item.value_state == "provided" and item.value is not None]
                state = values[0].value_state if values else "pending"
                note = next((item.source_note for item in values if item.source_note), "")
                payload[f"field:{definition.key}"] = "\n".join(provided)
                payload[f"field_state:{definition.key}"] = state
                payload[f"field_note:{definition.key}"] = note or ""
            rows_to_write.append(payload)
    if seed_rows:
        rows_to_write.extend(seed_rows)

    for row_index, payload in enumerate(rows_to_write, start=2):
        for key, column in header_map.items():
            value = payload.get(key, "")
            catalog.cell(row=row_index, column=column, value=value)
            catalog.cell(row=row_index, column=column).alignment = Alignment(
                vertical="top", wrap_text=True
            )

    _add_catalog_validations(catalog, header_map, lists)
    title_column = get_column_letter(header_map["title"])
    catalog.conditional_formatting.add(
        f"{title_column}2:{title_column}5000",
        FormulaRule(
            formula=[f'LEN(TRIM({title_column}2))=0'],
            fill=PatternFill("solid", fgColor="F4CCCC"),
        ),
    )
    widths: dict[int, float] = {}
    for key, column in header_map.items():
        if key in {"title", "source_note"} or key.startswith("field:") or key.startswith("field_note:"):
            widths[column] = 32
        elif key in {"source_url", "reference_code", "parent_local_id", "local_id"}:
            widths[column] = 24
        elif key in {"unit_id", "parent_unit_id"}:
            widths[column] = 18
        else:
            widths[column] = 18
    _set_widths(catalog, widths)
    for row in range(2, max(3, catalog.max_row + 1)):
        catalog.row_dimensions[row].height = 34

    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def write_catalog_template(
    destination: Path,
    session: Session | None,
    **kwargs: Any,
) -> Path:
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_bytes(export_catalog_template_bytes(session, **kwargs))
    return destination


def _load_source(source: Path | bytes | bytearray | BinaryIO):
    if isinstance(source, Path):
        return load_workbook(source, data_only=False)
    if isinstance(source, (bytes, bytearray)):
        return load_workbook(BytesIO(bytes(source)), data_only=False)
    return load_workbook(source, data_only=False)


def _metadata_from_sheet(sheet) -> dict[str, str]:
    result: dict[str, str] = {}
    for row in range(1, min(sheet.max_row, 30) + 1):
        key = _clean_text(sheet.cell(row=row, column=1).value)
        if key in {
            "schema_version",
            "template_name",
            "target_project_id",
            "project_name",
            "generated_at",
            "source_url",
            "source_retrieved_at",
            "source_note",
        }:
            result[key] = _clean_text(sheet.cell(row=row, column=2).value)
    return result


def _parse_template(source: Path | bytes | bytearray | BinaryIO, decisions: ProjectDecisions) -> _ParsedTemplate:
    issues: list[CatalogTemplateIssue] = []
    try:
        workbook = _load_source(source)
    except Exception as exc:
        return _ParsedTemplate(
            schema_version="",
            template_name="",
            target_project_id=None,
            structure_parent_keys={},
            rows=[],
            parse_issues=[
                CatalogTemplateIssue("error", "workbook_unreadable", "", None, None, str(exc))
            ],
        )
    missing = [name for name in TEMPLATE_SHEETS if name not in workbook.sheetnames]
    for name in missing:
        issues.append(
            CatalogTemplateIssue(
                "error", "missing_sheet", name, None, None, f"Falta la hoja obligatoria {name}."
            )
        )
    if missing:
        return _ParsedTemplate("", "", None, {}, [], issues)

    metadata = _metadata_from_sheet(workbook["INSTRUCCIONES"])
    schema_version = metadata.get("schema_version", "")
    template_name = metadata.get("template_name", "Plantilla sin nombre")
    target_project_id = metadata.get("target_project_id") or None

    structure_sheet = workbook["ESTRUCTURA"]
    structure_headers = {
        _column_key(structure_sheet.cell(row=1, column=column).value): column
        for column in range(1, structure_sheet.max_column + 1)
    }
    structure_parent_keys: dict[str, tuple[str, ...]] = {}
    required_structure = {"level_key", "template_parent_keys"}
    if not required_structure.issubset(structure_headers):
        issues.append(
            CatalogTemplateIssue(
                "error",
                "structure_headers",
                "ESTRUCTURA",
                1,
                None,
                "La hoja ESTRUCTURA no contiene las columnas técnicas requeridas.",
            )
        )
    else:
        for row in range(2, structure_sheet.max_row + 1):
            level_key = _clean_text(
                structure_sheet.cell(row=row, column=structure_headers["level_key"]).value
            )
            if not level_key:
                continue
            if level_key in structure_parent_keys:
                issues.append(
                    CatalogTemplateIssue(
                        "error",
                        "duplicate_structure_level",
                        "ESTRUCTURA",
                        row,
                        "level_key",
                        f"El nivel {level_key} aparece más de una vez.",
                    )
                )
                continue
            structure_parent_keys[level_key] = _split_keys(
                structure_sheet.cell(
                    row=row, column=structure_headers["template_parent_keys"]
                ).value
            )

    catalog_sheet = workbook["CATALOGO"]
    header_by_key: dict[str, int] = {}
    for column in range(1, catalog_sheet.max_column + 1):
        key = _column_key(catalog_sheet.cell(row=1, column=column).value)
        if key:
            header_by_key[key] = column
    required_catalog = {"local_id", "level_key", "title"}
    if not required_catalog.issubset(header_by_key):
        issues.append(
            CatalogTemplateIssue(
                "error",
                "catalog_headers",
                "CATALOGO",
                1,
                None,
                "La hoja CATALOGO no contiene local_id, level_key y title.",
            )
        )
        return _ParsedTemplate(
            schema_version,
            template_name,
            target_project_id,
            structure_parent_keys,
            [],
            issues,
        )

    field_map = {item.key: item for item in decisions.descriptive_fields if item.enabled and item.key != "reference_code"}
    rows: list[CatalogTemplateRow] = []
    for row_number in range(2, catalog_sheet.max_row + 1):
        raw = {
            key: catalog_sheet.cell(row=row_number, column=column).value
            for key, column in header_by_key.items()
        }
        if not any(_clean_text(value) for value in raw.values()):
            continue
        local_id = _clean_text(raw.get("local_id"))
        bool_value = _parse_bool(raw.get("completion_confirmed"))
        if bool_value is None:
            issues.append(
                CatalogTemplateIssue(
                    "error",
                    "invalid_boolean",
                    "CATALOGO",
                    row_number,
                    "completion_confirmed",
                    "La confirmación debe ser sí o no.",
                )
            )
            bool_value = False
        field_values: dict[str, dict[str, Any]] = {}
        source_url = _clean_text(raw.get("source_url")) or None
        row_note = _clean_text(raw.get("source_note")) or None
        for field_key, definition in field_map.items():
            value = raw.get(f"field:{field_key}")
            state = _clean_text(raw.get(f"field_state:{field_key}"))
            note = _clean_text(raw.get(f"field_note:{field_key}")) or None
            text_value = _clean_text(value)
            if not state:
                state = "provided" if text_value else "pending"
            values: list[Any] = []
            if text_value:
                values = [part.strip() for part in text_value.splitlines() if part.strip()]
                if not definition.repeatable and values:
                    values = values[:1]
            combined_note = note
            if not combined_note and source_url and (text_value or state != "pending"):
                combined_note = source_url
            field_values[field_key] = {
                "state": state,
                "values": values,
                "source_note": combined_note,
            }
        rows.append(
            CatalogTemplateRow(
                row_number=row_number,
                local_id=local_id,
                unit_id=_clean_text(raw.get("unit_id")) or None,
                parent_local_id=_clean_text(raw.get("parent_local_id")) or None,
                parent_unit_id=_clean_text(raw.get("parent_unit_id")) or None,
                action=_clean_text(raw.get("action")).casefold(),
                level_key=_clean_text(raw.get("level_key")),
                reference_code=_clean_text(raw.get("reference_code")) or None,
                title=_clean_text(raw.get("title")),
                registration_status=_clean_text(raw.get("registration_status")) or "incomplete",
                completion_confirmed=bool_value,
                source_url=source_url,
                source_note=row_note,
                field_values=field_values,
            )
        )
    return _ParsedTemplate(
        schema_version,
        template_name,
        target_project_id,
        structure_parent_keys,
        rows,
        issues,
    )


def _issue(
    issues: list[CatalogTemplateIssue],
    severity: str,
    code: str,
    sheet: str,
    row: int | None,
    column: str | None,
    message: str,
) -> None:
    issues.append(CatalogTemplateIssue(severity, code, sheet, row, column, message))


def _effective_action(row: CatalogTemplateRow) -> str:
    if row.action:
        return row.action
    return "actualizar" if row.unit_id else "crear"


def _validate_field_value(definition, payload: dict[str, Any]) -> str | None:
    state = payload.get("state")
    values = payload.get("values", [])
    if state not in {"provided", "no_information", "not_applicable", "pending"}:
        return f"Estado de campo inválido: {state}"
    if state == "provided" and not values:
        return "El estado provided requiere un valor."
    if state != "provided" and values:
        return f"El estado {state} no admite valores."
    if not values:
        return None
    for value in values:
        text = str(value).strip()
        try:
            if definition.data_type == "integer":
                int(text)
            elif definition.data_type == "number":
                float(text.replace(",", "."))
            elif definition.data_type == "boolean" and _parse_bool(text) is None:
                return f"Valor booleano inválido: {text}"
            elif definition.data_type == "date":
                date.fromisoformat(text)
            elif definition.data_type == "date_range":
                if not text:
                    return "El rango de fechas no puede estar vacío."
        except ValueError:
            return f"Valor inválido para tipo {definition.data_type}: {text}"
    return None


def validate_catalog_template(
    session: Session,
    *,
    decisions: ProjectDecisions,
    project_id: str,
    source: Path | bytes | bytearray | BinaryIO,
) -> CatalogTemplateReport:
    parsed = _parse_template(source, decisions)
    issues = list(parsed.parse_issues)
    if parsed.schema_version != TEMPLATE_SCHEMA_VERSION:
        _issue(
            issues,
            "error",
            "schema_version",
            "INSTRUCCIONES",
            None,
            "schema_version",
            f"Versión de plantilla incompatible: {parsed.schema_version or 'vacía'}. Se requiere {TEMPLATE_SCHEMA_VERSION}.",
        )
    if parsed.target_project_id not in {None, "", "*", project_id}:
        _issue(
            issues,
            "error",
            "target_project",
            "INSTRUCCIONES",
            None,
            "target_project_id",
            f"La plantilla está dirigida al proyecto {parsed.target_project_id}, no a {project_id}.",
        )

    level_map = {item.key: item for item in decisions.archival_levels if item.enabled}
    for level_key, definition in level_map.items():
        if level_key not in parsed.structure_parent_keys:
            _issue(
                issues,
                "error",
                "missing_structure_level",
                "ESTRUCTURA",
                None,
                "level_key",
                f"Falta el nivel habilitado {level_key}.",
            )
            continue
        template_parents = set(parsed.structure_parent_keys[level_key])
        project_parents = set(definition.parent_keys)
        unknown = sorted(template_parents - project_parents)
        if unknown:
            _issue(
                issues,
                "error",
                "structure_broadens_project",
                "ESTRUCTURA",
                None,
                "template_parent_keys",
                f"{level_key} intenta admitir padres no permitidos por el proyecto: {', '.join(unknown)}.",
            )
    for level_key in sorted(set(parsed.structure_parent_keys) - set(level_map)):
        _issue(
            issues,
            "error",
            "unknown_structure_level",
            "ESTRUCTURA",
            None,
            "level_key",
            f"Nivel desconocido en la plantilla: {level_key}.",
        )

    local_rows: dict[str, CatalogTemplateRow] = {}
    existing_units = {
        row.id: row
        for row in session.scalars(
            select(ArchivalUnit).where(ArchivalUnit.project_id == project_id)
        ).all()
    }
    field_map = {
        item.key: item
        for item in decisions.descriptive_fields
        if item.enabled and item.key != "reference_code"
    }
    create_count = update_count = skip_count = 0
    for row in parsed.rows:
        action = _effective_action(row)
        if action not in {"crear", "actualizar", "omitir"}:
            _issue(
                issues,
                "error",
                "invalid_action",
                "CATALOGO",
                row.row_number,
                "action",
                f"Acción inválida: {row.action or action}.",
            )
        if action == "crear":
            create_count += 1
        elif action == "actualizar":
            update_count += 1
        elif action == "omitir":
            skip_count += 1
        if not row.local_id:
            _issue(
                issues,
                "error",
                "missing_local_id",
                "CATALOGO",
                row.row_number,
                "local_id",
                "Cada fila necesita un ID local.",
            )
        elif row.local_id in local_rows:
            _issue(
                issues,
                "error",
                "duplicate_local_id",
                "CATALOGO",
                row.row_number,
                "local_id",
                f"ID local repetido: {row.local_id}.",
            )
        else:
            local_rows[row.local_id] = row
        if row.level_key not in level_map:
            _issue(
                issues,
                "error",
                "unknown_level",
                "CATALOGO",
                row.row_number,
                "level_key",
                f"Nivel desconocido: {row.level_key or 'vacío'}.",
            )
        if not row.title:
            _issue(
                issues,
                "error",
                "missing_title",
                "CATALOGO",
                row.row_number,
                "title",
                "El título es obligatorio.",
            )
        if row.unit_id and row.unit_id not in existing_units:
            _issue(
                issues,
                "error",
                "unknown_unit_id",
                "CATALOGO",
                row.row_number,
                "unit_id",
                f"La unidad existente {row.unit_id} no pertenece a este proyecto.",
            )
        if action == "crear" and row.unit_id:
            _issue(
                issues,
                "error",
                "create_with_unit_id",
                "CATALOGO",
                row.row_number,
                "unit_id",
                "Una fila marcada crear no debe tener ID de unidad existente.",
            )
        if action == "actualizar" and not row.unit_id:
            _issue(
                issues,
                "error",
                "update_without_unit_id",
                "CATALOGO",
                row.row_number,
                "unit_id",
                "Una fila marcada actualizar necesita ID de unidad existente.",
            )
        if row.parent_local_id and row.parent_unit_id:
            _issue(
                issues,
                "error",
                "two_parents",
                "CATALOGO",
                row.row_number,
                "parent_local_id",
                "Usá ID padre local o ID padre existente, no ambos.",
            )
        if row.parent_unit_id and row.parent_unit_id not in existing_units:
            _issue(
                issues,
                "error",
                "unknown_parent_unit",
                "CATALOGO",
                row.row_number,
                "parent_unit_id",
                f"La unidad padre {row.parent_unit_id} no existe en este proyecto.",
            )
        if row.registration_status not in REGISTRATION_STATUSES:
            _issue(
                issues,
                "error",
                "invalid_registration_status",
                "CATALOGO",
                row.row_number,
                "registration_status",
                f"Estado de registro inválido: {row.registration_status}.",
            )
        if row.registration_status == "complete" and not row.completion_confirmed:
            _issue(
                issues,
                "error",
                "complete_without_confirmation",
                "CATALOGO",
                row.row_number,
                "completion_confirmed",
                "Un registro complete requiere confirmación manual.",
            )
        for field_key, payload in row.field_values.items():
            definition = field_map[field_key]
            if row.level_key in level_map and (
                "all" not in definition.applies_to_levels
                and row.level_key not in definition.applies_to_levels
            ):
                if payload.get("state") != "pending" or payload.get("values"):
                    _issue(
                        issues,
                        "error",
                        "field_not_applicable",
                        "CATALOGO",
                        row.row_number,
                        f"field:{field_key}",
                        f"{definition.label} no se aplica al nivel {row.level_key}.",
                    )
                continue
            field_error = _validate_field_value(definition, payload)
            if field_error:
                _issue(
                    issues,
                    "error",
                    "invalid_field_value",
                    "CATALOGO",
                    row.row_number,
                    f"field:{field_key}",
                    f"{definition.label}: {field_error}",
                )
            if definition.required and payload.get("state") != "provided":
                _issue(
                    issues,
                    "error",
                    "required_field",
                    "CATALOGO",
                    row.row_number,
                    f"field:{field_key}",
                    f"El campo obligatorio {definition.label} debe estar informado.",
                )

    for row in parsed.rows:
        if _effective_action(row) == "omitir" or row.level_key not in level_map:
            continue
        parent_level: str | None = None
        if row.parent_local_id:
            parent_row = local_rows.get(row.parent_local_id)
            if parent_row is None:
                _issue(
                    issues,
                    "error",
                    "missing_parent_local",
                    "CATALOGO",
                    row.row_number,
                    "parent_local_id",
                    f"No existe el ID padre local {row.parent_local_id}.",
                )
            elif _effective_action(parent_row) == "omitir" and not parent_row.unit_id:
                _issue(
                    issues,
                    "error",
                    "skipped_new_parent",
                    "CATALOGO",
                    row.row_number,
                    "parent_local_id",
                    "La unidad padre se omite y todavía no existe en el proyecto.",
                )
            else:
                parent_level = parent_row.level_key
        elif row.parent_unit_id:
            parent_level = existing_units[row.parent_unit_id].level_key
        allowed = set(parsed.structure_parent_keys.get(row.level_key, ()))
        if parent_level is None:
            if allowed:
                _issue(
                    issues,
                    "error",
                    "missing_required_parent",
                    "CATALOGO",
                    row.row_number,
                    "parent_local_id",
                    f"El nivel {row.level_key} requiere padre de tipo: {', '.join(sorted(allowed))}.",
                )
        elif parent_level not in allowed:
            _issue(
                issues,
                "error",
                "invalid_parent_level",
                "CATALOGO",
                row.row_number,
                "parent_local_id",
                f"{parent_level} no puede contener una unidad de nivel {row.level_key} según la plantilla.",
            )

    visiting: set[str] = set()
    visited: set[str] = set()

    def visit(local_id: str) -> None:
        if local_id in visited:
            return
        if local_id in visiting:
            cycle_row = local_rows[local_id]
            _issue(
                issues,
                "error",
                "catalog_cycle",
                "CATALOGO",
                cycle_row.row_number,
                "parent_local_id",
                f"Ciclo detectado cerca de {local_id}.",
            )
            return
        visiting.add(local_id)
        row = local_rows[local_id]
        if row.parent_local_id and row.parent_local_id in local_rows:
            visit(row.parent_local_id)
        visiting.remove(local_id)
        visited.add(local_id)

    for local_id in local_rows:
        visit(local_id)

    valid = not any(item.severity == "error" for item in issues)
    return CatalogTemplateReport(
        schema_version=parsed.schema_version,
        template_name=parsed.template_name,
        target_project_id=parsed.target_project_id,
        valid=valid,
        rows=parsed.rows,
        structure_parent_keys=parsed.structure_parent_keys,
        issues=issues,
        create_count=create_count,
        update_count=update_count,
        skip_count=skip_count,
    )


def _normalized_existing_fields(session: Session, unit_id: str) -> dict[str, dict[str, Any]]:
    result: dict[str, dict[str, Any]] = {}
    for row in archival_field_rows(session, unit_id):
        payload = result.setdefault(
            row.field_key,
            {"state": row.value_state, "values": [], "source_note": row.source_note},
        )
        if row.value_state == "provided" and row.value is not None:
            payload["values"].append(row.value)
        if not payload.get("source_note") and row.source_note:
            payload["source_note"] = row.source_note
    return result


def _applicable_payload(row: CatalogTemplateRow, decisions: ProjectDecisions) -> dict[str, dict[str, Any]]:
    definitions = {
        item.key: item
        for item in decisions.descriptive_fields
        if item.enabled and item.key != "reference_code"
    }
    return {
        key: payload
        for key, payload in row.field_values.items()
        if key in definitions
        and ("all" in definitions[key].applies_to_levels or row.level_key in definitions[key].applies_to_levels)
    }


def _fields_equal(left: dict[str, dict[str, Any]], right: dict[str, dict[str, Any]]) -> bool:
    keys = set(left) | set(right)
    for key in keys:
        a = left.get(key, {"state": "pending", "values": [], "source_note": None})
        b = right.get(key, {"state": "pending", "values": [], "source_note": None})
        if a.get("state", "pending") != b.get("state", "pending"):
            return False
        if [str(item) for item in a.get("values", [])] != [str(item) for item in b.get("values", [])]:
            return False
        if (a.get("source_note") or None) != (b.get("source_note") or None):
            return False
    return True


def apply_catalog_template(
    session: Session,
    *,
    decisions: ProjectDecisions,
    project_id: str,
    source: Path | bytes | bytearray | BinaryIO,
    changed_by: str,
    note: str | None = None,
) -> CatalogTemplateApplyResult:
    report = validate_catalog_template(
        session, decisions=decisions, project_id=project_id, source=source
    )
    if not report.valid:
        first = next(item for item in report.issues if item.severity == "error")
        location = f"{first.sheet} fila {first.row}" if first.row else first.sheet
        raise ValueError(f"La plantilla contiene errores ({location}): {first.message}")

    actor = changed_by.strip() or "local_user"
    ensure_project(session, decisions)
    local_to_unit_id = {
        row.local_id: row.unit_id
        for row in report.rows
        if row.unit_id
    }
    created = updated = moved = unchanged = skipped = 0
    pending = [row for row in report.rows if _effective_action(row) != "omitir"]
    skipped = sum(_effective_action(row) == "omitir" for row in report.rows)
    processed: set[str] = set()
    while pending:
        progressed = False
        for row in list(pending):
            if row.parent_local_id and row.parent_local_id not in local_to_unit_id:
                continue
            parent_id = (
                local_to_unit_id.get(row.parent_local_id or "")
                if row.parent_local_id
                else row.parent_unit_id
            )
            action = _effective_action(row)
            row_note_parts = [part for part in [note, row.source_note] if part]
            if row.source_url:
                row_note_parts.append(f"Fuente: {row.source_url}")
            row_note = " · ".join(row_note_parts) or f"Importación desde {report.template_name}"
            payload = _applicable_payload(row, decisions)
            if action == "crear":
                unit = create_archival_unit(
                    session,
                    decisions=decisions,
                    project_id=project_id,
                    parent_id=parent_id,
                    level_key=row.level_key,
                    title=row.title,
                    created_by=actor,
                    reference_code=row.reference_code,
                    note=row_note,
                )
                local_to_unit_id[row.local_id] = unit.id
                created += 1
                if (
                    payload
                    or row.registration_status != "incomplete"
                    or row.completion_confirmed
                ):
                    update_archival_unit(
                        session,
                        decisions=decisions,
                        unit_id=unit.id,
                        changed_by=actor,
                        title=row.title,
                        reference_code=row.reference_code,
                        registration_status=row.registration_status,
                        completion_confirmed=row.completion_confirmed,
                        field_values=payload,
                        note=row_note,
                    )
            else:
                unit = session.get(ArchivalUnit, row.unit_id)
                if unit is None:
                    raise ValueError(f"La unidad {row.unit_id} dejó de existir durante la importación.")
                did_move = unit.parent_id != parent_id
                if did_move:
                    move_archival_unit(
                        session,
                        decisions=decisions,
                        unit_id=unit.id,
                        new_parent_id=parent_id,
                        changed_by=actor,
                        note=row_note,
                    )
                    moved += 1
                existing_fields = _normalized_existing_fields(session, unit.id)
                needs_update = (
                    unit.title != row.title
                    or unit.reference_code != row.reference_code
                    or unit.registration_status != row.registration_status
                    or bool(unit.completion_confirmed) != row.completion_confirmed
                    or not _fields_equal(existing_fields, payload)
                )
                if needs_update:
                    update_archival_unit(
                        session,
                        decisions=decisions,
                        unit_id=unit.id,
                        changed_by=actor,
                        title=row.title,
                        reference_code=row.reference_code,
                        registration_status=row.registration_status,
                        completion_confirmed=row.completion_confirmed,
                        field_values=payload,
                        note=row_note,
                    )
                    updated += 1
                elif not did_move:
                    unchanged += 1
                local_to_unit_id[row.local_id] = unit.id
            processed.add(row.local_id)
            pending.remove(row)
            progressed = True
        if not progressed:
            unresolved = ", ".join(row.local_id for row in pending[:5])
            raise ValueError(f"No se pudo resolver el orden jerárquico de la importación: {unresolved}")
    session.flush()
    return CatalogTemplateApplyResult(
        created=created,
        updated=updated,
        moved=moved,
        unchanged=unchanged,
        skipped=skipped,
        local_to_unit_id={key: value for key, value in local_to_unit_id.items() if value},
    )
