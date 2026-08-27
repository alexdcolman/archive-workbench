from __future__ import annotations

from io import BytesIO
import json
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import func, select

from archive_workbench.catalog import ensure_project
from archive_workbench.catalog_management import create_archival_unit
from archive_workbench.catalog_templates import (
    TEMPLATE_SCHEMA_VERSION,
    apply_catalog_template,
    export_catalog_template_bytes,
    validate_catalog_template,
)
from archive_workbench.db import create_sqlite_engine, database_path, session_scope, upgrade_database
from archive_workbench.db.models import ArchivalFieldValue, ArchivalUnit, ArchivalUnitRevision, Project
from archive_workbench.decisions import load_decisions


def _setup(tmp_path: Path):
    root = tmp_path / "project"
    upgrade_database(root)
    decisions = load_decisions(Path(__file__).parents[1] / "config" / "decisions.yaml")
    engine = create_sqlite_engine(database_path(root))
    with session_scope(engine) as session:
        ensure_project(session, decisions)
    return root, decisions, engine


def _header_map(sheet) -> dict[str, int]:
    result: dict[str, int] = {}
    for column in range(1, sheet.max_column + 1):
        text = str(sheet.cell(row=1, column=column).value or "")
        if "[" in text and text.endswith("]"):
            result[text.rsplit("[", 1)[1][:-1]] = column
    return result


def test_export_template_contains_required_sheets_and_configurable_structure(tmp_path: Path) -> None:
    _, decisions, engine = _setup(tmp_path)
    try:
        with session_scope(engine) as session:
            content = export_catalog_template_bytes(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                include_catalog=False,
                template_name="Control CAT-01",
                structure_parent_overrides={"documento": ["caja", "legajo", "tomo"]},
            )
    finally:
        engine.dispose()

    workbook = load_workbook(BytesIO(content), data_only=False)
    assert workbook.sheetnames == ["INSTRUCCIONES", "ESTRUCTURA", "CATALOGO", "LISTAS"]
    assert workbook["LISTAS"].sheet_state == "hidden"
    instructions_text = " ".join(
        str(workbook["INSTRUCCIONES"].cell(row=row, column=2).value or "")
        for row in range(1, workbook["INSTRUCCIONES"].max_row + 1)
    )
    assert "LISTAS es una hoja técnica oculta" in instructions_text
    assert "Archive Workbench revisa toda la plantilla" in instructions_text
    assert any(
        workbook["INSTRUCCIONES"].cell(row=row, column=5).value == "Ejemplo"
        for row in range(1, workbook["INSTRUCCIONES"].max_row + 1)
    )
    assert any(
        str(workbook["INSTRUCCIONES"].cell(row=row, column=5).value or "").strip()
        for row in range(1, workbook["INSTRUCCIONES"].max_row + 1)
        if workbook["INSTRUCCIONES"].cell(row=row, column=3).value == "title"
    )
    metadata = {
        workbook["INSTRUCCIONES"].cell(row=row, column=1).value:
        workbook["INSTRUCCIONES"].cell(row=row, column=2).value
        for row in range(1, 15)
    }
    assert metadata["schema_version"] == TEMPLATE_SCHEMA_VERSION
    structure = workbook["ESTRUCTURA"]
    headers = _header_map(structure)
    document_row = next(
        row
        for row in range(2, structure.max_row + 1)
        if structure.cell(row=row, column=headers["level_key"]).value == "documento"
    )
    assert structure.cell(
        row=document_row, column=headers["template_parent_keys"]
    ).value == "caja, legajo, tomo"
    catalog_headers = _header_map(workbook["CATALOGO"])
    assert {"local_id", "parent_local_id", "level_key", "title"}.issubset(catalog_headers)
    assert "field:scope_content" in catalog_headers


def test_validation_rejects_duplicate_ids_cycles_and_invalid_template_parent(tmp_path: Path) -> None:
    _, decisions, engine = _setup(tmp_path)
    try:
        seed = [
            {
                "local_id": "fondo",
                "parent_local_id": "documento",
                "level_key": "fondo",
                "title": "Fondo",
                "registration_status": "provisional",
            },
            {
                "local_id": "documento",
                "parent_local_id": "fondo",
                "level_key": "documento",
                "title": "Documento",
                "registration_status": "provisional",
            },
            {
                "local_id": "documento",
                "parent_local_id": "fondo",
                "level_key": "documento",
                "title": "Duplicado",
                "registration_status": "provisional",
            },
        ]
        with session_scope(engine) as session:
            content = export_catalog_template_bytes(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                seed_rows=seed,
                structure_parent_overrides={
                    "archivo": [],
                    "fondo": ["archivo"],
                    "documento": ["legajo"],
                },
            )
            report = validate_catalog_template(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                source=content,
            )
    finally:
        engine.dispose()
    codes = {issue.code for issue in report.issues}
    assert not report.valid
    assert "duplicate_local_id" in codes
    assert "catalog_cycle" in codes
    assert "invalid_parent_level" in codes or "missing_required_parent" in codes


def test_apply_template_creates_hierarchy_and_preserves_field_provenance(tmp_path: Path) -> None:
    _, decisions, engine = _setup(tmp_path)
    try:
        rows = [
            {
                "local_id": "archivo",
                "level_key": "archivo",
                "title": "Archivo de prueba",
                "registration_status": "provisional",
            },
            {
                "local_id": "fondo",
                "parent_local_id": "archivo",
                "level_key": "fondo",
                "reference_code": "AR.TEST",
                "title": "Fondo de prueba",
                "registration_status": "provisional",
            },
            {
                "local_id": "serie",
                "parent_local_id": "fondo",
                "level_key": "serie",
                "title": "Serie de informes",
                "registration_status": "provisional",
                "source_url": "https://example.org/cuadro",
                "field:scope_content": "Informes y correspondencia",
                "field_state:scope_content": "provided",
            },
        ]
        with session_scope(engine) as session:
            content = export_catalog_template_bytes(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                seed_rows=rows,
            )
            report = validate_catalog_template(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                source=content,
            )
            before = session.scalar(select(func.count()).select_from(ArchivalUnit))
            assert report.valid
            assert before == 0
        with session_scope(engine) as session:
            result = apply_catalog_template(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                source=content,
                changed_by="Alex",
            )
            assert result.created == 3
        with session_scope(engine) as session:
            units = session.scalars(select(ArchivalUnit).order_by(ArchivalUnit.title)).all()
            by_title = {unit.title: unit for unit in units}
            assert by_title["Fondo de prueba"].parent_id == by_title["Archivo de prueba"].id
            assert by_title["Serie de informes"].parent_id == by_title["Fondo de prueba"].id
            field = session.scalar(
                select(ArchivalFieldValue).where(
                    ArchivalFieldValue.archival_unit_id == by_title["Serie de informes"].id,
                    ArchivalFieldValue.field_key == "scope_content",
                )
            )
            assert field is not None
            assert field.value_json == "Informes y correspondencia"
            assert field.source_note == "https://example.org/cuadro"
            revision = session.scalar(
                select(ArchivalUnitRevision)
                .where(ArchivalUnitRevision.archival_unit_id == by_title["Serie de informes"].id)
                .order_by(ArchivalUnitRevision.revision_number.desc())
            )
            assert revision is not None
            assert "Fuente: https://example.org/cuadro" in (revision.note or "")
    finally:
        engine.dispose()



def test_apply_template_registers_project_in_fresh_database(tmp_path: Path) -> None:
    root = tmp_path / "fresh_project"
    upgrade_database(root)
    decisions = load_decisions(Path(__file__).parents[1] / "config" / "decisions.yaml")
    engine = create_sqlite_engine(database_path(root))
    try:
        rows = [
            {
                "local_id": "archivo",
                "level_key": "archivo",
                "title": "Archivo inicial",
                "registration_status": "incomplete",
            },
            {
                "local_id": "fondo",
                "parent_local_id": "archivo",
                "level_key": "fondo",
                "title": "Fondo inicial",
                "registration_status": "incomplete",
            },
        ]
        with session_scope(engine) as session:
            assert session.get(Project, decisions.project_id) is None
            content = export_catalog_template_bytes(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                seed_rows=rows,
            )
            report = validate_catalog_template(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                source=content,
            )
            assert report.valid
            assert session.get(Project, decisions.project_id) is None

        with session_scope(engine) as session:
            result = apply_catalog_template(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                source=content,
                changed_by="Alex",
            )
            assert result.created == 2

        with session_scope(engine) as session:
            project = session.get(Project, decisions.project_id)
            assert project is not None
            assert project.name == decisions.project_name
            assert session.scalar(select(func.count()).select_from(ArchivalUnit)) == 2
    finally:
        engine.dispose()


def test_apply_template_resolves_existing_omitted_parent_local_id(tmp_path: Path) -> None:
    _, decisions, engine = _setup(tmp_path)
    try:
        with session_scope(engine) as session:
            archivo = create_archival_unit(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                parent_id=None,
                level_key="archivo",
                title="Archivo",
                created_by="Alex",
            )
            fondo = create_archival_unit(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                parent_id=archivo.id,
                level_key="fondo",
                title="Fondo",
                created_by="Alex",
            )

        with session_scope(engine) as session:
            content = export_catalog_template_bytes(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                include_catalog=True,
            )

        workbook = load_workbook(BytesIO(content), data_only=False)
        sheet = workbook["CATALOGO"]
        headers = _header_map(sheet)
        by_level = {
            sheet.cell(row=row, column=headers["level_key"]).value: row
            for row in range(2, sheet.max_row + 1)
        }
        archivo_row = by_level["archivo"]
        fondo_row = by_level["fondo"]
        sheet.cell(row=archivo_row, column=headers["action"]).value = "omitir"
        sheet.cell(row=fondo_row, column=headers["title"]).value = "Fondo actualizado"
        buffer = BytesIO()
        workbook.save(buffer)
        content = buffer.getvalue()

        with session_scope(engine) as session:
            report = validate_catalog_template(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                source=content,
            )
            assert report.valid
            assert report.skip_count == 1
            assert report.update_count == 1

        with session_scope(engine) as session:
            result = apply_catalog_template(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                source=content,
                changed_by="Alex",
            )
            assert result.created == 0
            assert result.updated == 1
            assert result.skipped == 1

        with session_scope(engine) as session:
            refreshed_archivo = session.get(ArchivalUnit, archivo.id)
            refreshed_fondo = session.get(ArchivalUnit, fondo.id)
            assert refreshed_archivo is not None
            assert refreshed_archivo.title == "Archivo"
            assert refreshed_fondo is not None
            assert refreshed_fondo.title == "Fondo actualizado"
            assert refreshed_fondo.parent_id == archivo.id
    finally:
        engine.dispose()

def test_exported_existing_catalog_roundtrip_is_valid_and_unchanged(tmp_path: Path) -> None:
    _, decisions, engine = _setup(tmp_path)
    try:
        with session_scope(engine) as session:
            archivo = create_archival_unit(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                parent_id=None,
                level_key="archivo",
                title="Archivo",
                created_by="Alex",
            )
            create_archival_unit(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                parent_id=archivo.id,
                level_key="fondo",
                title="Fondo",
                created_by="Alex",
            )
        with session_scope(engine) as session:
            content = export_catalog_template_bytes(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                include_catalog=True,
            )
            report = validate_catalog_template(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                source=content,
            )
            assert report.valid
        with session_scope(engine) as session:
            result = apply_catalog_template(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                source=content,
                changed_by="Alex",
            )
            assert result.created == 0
            assert result.updated == 0
            assert result.moved == 0
            assert result.unchanged == 2
    finally:
        engine.dispose()


def test_dippba_public_seed_and_example_are_traceable_and_valid(tmp_path: Path) -> None:
    root = Path(__file__).parents[1]
    seed_path = root / "config" / "catalog_templates" / "dippba_public_seed.json"
    example_path = root / "examples" / "plantilla_catalogo_dippba.xlsx"
    seed = json.loads(seed_path.read_text(encoding="utf-8"))
    assert len(seed["rows"]) == 155
    assert seed["source_url"].startswith("https://www.comisionporlamemoria.org/")
    assert any(row["title"] == "MESA A" for row in seed["rows"])
    assert any(row["title"] == "FACTOR POLITICO" for row in seed["rows"])
    assert any("elipsis" in row.get("source_note", "") for row in seed["rows"])
    assert example_path.is_file()

    _, decisions, engine = _setup(tmp_path)
    try:
        with session_scope(engine) as session:
            report = validate_catalog_template(
                session,
                decisions=decisions,
                project_id=decisions.project_id,
                source=example_path,
            )
            assert report.valid
            assert report.create_count == 155
    finally:
        engine.dispose()
